"""
KAP Pay Alım Satım Bildirimi Scraper
- Liste:  POST /tr/api/disclosure/list/main
- Detay:  GET  /tr/api/BildirimPdf/{discIndex}  → PDF
- Parse:  pdfplumber ile tablo çıkarma

Kurulum:
    pip install requests pandas openpyxl beautifulsoup4 lxml pdfplumber
"""

import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import re, os, sys, time, json, io

BASE_URL = "https://www.kap.org.tr"
HEADERS  = {
    "User-Agent":    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                     "(KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36",
    "Accept":        "*/*",
    "Accept-Language":"tr",
    "Content-Type":  "application/json",
    "Origin":        BASE_URL,
    "Referer":       f"{BASE_URL}/tr",
}


# ─── LİSTE API ───────────────────────────────────────────

def make_session():
    s = requests.Session()
    s.headers.update(HEADERS)
    try:
        s.get(f"{BASE_URL}/tr", timeout=15)
    except Exception as e:
        print(f"  Session uyarı: {e}")
    return s


def fetch_main_api(session, start_date: date, end_date: date) -> list:
    payload = {
        "fromDate":    start_date.strftime("%d.%m.%Y"),
        "toDate":      end_date.strftime("%d.%m.%Y"),
        "memberTypes": ["IGS", "DDK"],
    }
    try:
        r = session.post(f"{BASE_URL}/tr/api/disclosure/list/main",
                         json=payload, timeout=30)
        if r.status_code == 200:
            data = r.json()
            if isinstance(data, list) and data:
                print(f"  ✓ API: {len(data)} kayıt")
                return data
        print(f"  API HTTP {r.status_code}")
    except Exception as e:
        print(f"  API hata: {e}")
    return []


def filter_pay_alim_satim(items: list) -> list:
    kw = ["pay alım satım", "pay alim satim"]
    out = []
    for item in items:
        b = item.get("disclosureBasic", item)
        t = " ".join([
            b.get("title","") or "", b.get("summary","") or ""
        ]).lower().replace("ı","i")
        if any(k.replace("ı","i") in t for k in kw):
            out.append(item)
    return out


def normalize_item(item: dict) -> dict:
    b     = item.get("disclosureBasic", item)
    idx   = b.get("disclosureIndex") or ""
    link  = f"{BASE_URL}/tr/Bildirim/{idx}" if idx else ""

    rs_raw = b.get("relatedStocks") or ""
    ilgili = _clean_related(rs_raw)

    stock_code = (b.get("stockCode") or "").strip()
    company = b.get("companyTitle") or ""

    return {
        "no":            str(idx),
        "tarih":         b.get("publishDate") or "",
        "kod":           stock_code,
        "sirket":        company,
        "konu":          b.get("title") or "",
        "ozet":          b.get("summary") or "",
        "link":          link,
        "disc_index":    str(idx),
        "ilgili_sirket": ilgili,
    }


def _clean_related(rs) -> str:
    if not rs:
        return ""
    if isinstance(rs, list):
        parts = []
        for r in rs:
            if isinstance(r, dict):
                c = r.get("stockCode","") or r.get("code","")
            else:
                c = str(r)
            c = c.strip().replace("[","").replace("]","")
            if 2 <= len(c) <= 10:
                parts.append(c)
        return ", ".join(parts)
    clean = str(rs).replace("[","").replace("]","").strip()
    parts = [p.strip() for p in clean.split() if 2 <= len(p.strip()) <= 10]
    return ", ".join(parts)


# ─── PDF PARSE ────────────────────────────────────────────

def _parse_kap_pdf(pdf_bytes: bytes, log_fn=print) -> dict:
    """
    KAP Pay Alım Satım PDF'ini parse et.

    KAP'ın standart şablon tablosu (10 kolon, test edildi):
    Row 0 (header): İşlem Tarihi | Alım Nominal | Satım Nominal | Net Nominal |
                     Gün Başı Nominal | Gün Sonu Nominal |
                     Sermaye Oranı Gün Başı | Oy Hakları Gün Başı |
                     Sermaye Oranı Gün Sonu | Oy Hakları Gün Sonu
    Row 1+ (veri):   01/04/2026 | 12.000 | 0 | 12.000 | 336.680 | 348.680 | 0,156 | 0,156 | 0,161 | 0,161

    Diğer format (portföy şirketleri — serbest form tablo):
    İşlem Tarihi | Niteliği | Nominal | Fiyat | İşlem Tutarı |
    Öncesi Nominal | Öncesi Sermaye | Sonrası Nominal | Sonrası Sermaye
    """
    result = {
        "islem_tarihi": "",
        "alim_toplam_nominal": "",
        "satim_toplam_nominal": "",
        "net_nominal": "",
        "gun_basi_nominal": "",
        "gun_sonu_nominal": "",
        "sermaye_orani_gun_basi": "",
        "oy_haklari_gun_basi": "",
        "sermaye_orani_gun_sonu": "",
        "oy_haklari_gun_sonu": "",
        "fiyat": "",
        "sirket_adi": "",
    }

    try:
        import pdfplumber
    except ImportError:
        log_fn("    ⚠ pdfplumber kurulu değil")
        return result

    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            all_text = ""
            all_tables = []
            for page in pdf.pages:
                text = page.extract_text() or ""
                all_text += text + "\n"
                tables = page.extract_tables()
                for t in tables:
                    all_tables.append(t)

            log_fn(f"    📊 {len(pdf.pages)} sayfa, {len(all_tables)} tablo")

            # ── Şirket adı (sayfa 1'deki bilgi tablolarından) ──
            result["sirket_adi"] = _extract_company(all_tables, all_text)

            # ── Fiyat (metin içinden) ──
            result["fiyat"] = _extract_price(all_text)

            # ── Ana veri tablosunu bul ──
            for table in all_tables:
                if not table or len(table) < 2:
                    continue

                # Veri satırı bul: tarih formatı içeren satır
                for row_idx, row in enumerate(table):
                    if not row:
                        continue
                    # Satırdaki herhangi bir hücrede tarih var mı?
                    row_str = " ".join(str(c or "") for c in row)
                    if not re.search(r'\d{2}[/\.]\d{2}[/\.]\d{4}', row_str):
                        continue

                    # Kaç kolon var?
                    ncols = len(row)
                    log_fn(f"    📋 Veri satırı: {ncols} kolon → {[str(c)[:15] for c in row]}")

                    if ncols >= 9:
                        # ── KAP standart şablon: 10 (veya 9) kolon ──
                        # Doğrudan pozisyonel eşle
                        parsed = _map_standard_row(row)
                        if parsed.get("islem_tarihi"):
                            result.update(parsed)
                            log_fn(f"    ✓ Standart tablo: {parsed['islem_tarihi']}")
                            return result

                    elif ncols >= 5:
                        # ── Portföy şirketi formatı ──
                        parsed = _map_portfolio_row(row)
                        if parsed.get("islem_tarihi"):
                            result.update(parsed)
                            log_fn(f"    ✓ Portföy tablo: {parsed['islem_tarihi']}")
                            return result

            # ── Tablo bulunamadı → metin parse ──
            text_result = _parse_text_fallback(all_text)
            if text_result:
                result.update(text_result)
                log_fn(f"    ✓ Metin parse: {text_result.get('islem_tarihi','?')}")

    except Exception as e:
        log_fn(f"    ✗ PDF parse hata: {e}")

    return result


def _map_standard_row(row: list) -> dict:
    """
    KAP standart 10-kolon tablosu (doğrudan dashboard kolonlarına eşlenir).
    [0] İşlem Tarihi       → islem_tarihi
    [1] Alım Nominal       → alim_toplam_nominal
    [2] Satım Nominal      → satim_toplam_nominal
    [3] Net Nominal         → net_nominal
    [4] Gün Başı Nominal   → gun_basi_nominal
    [5] Gün Sonu Nominal   → gun_sonu_nominal
    [6] Sermaye Gün Başı % → sermaye_orani_gun_basi
    [7] Oy Hakkı Gün Başı  → oy_haklari_gun_basi
    [8] Sermaye Gün Sonu % → sermaye_orani_gun_sonu
    [9] Oy Hakkı Gün Sonu  → oy_haklari_gun_sonu
    """
    def g(idx):
        if idx < len(row) and row[idx] is not None:
            return str(row[idx]).strip().replace("\n"," ")
        return ""

    tarih = g(0)
    # Tarih formatını düzelt: 01/04/2026 → 01.04.2026
    tarih = tarih.replace("/", ".")

    return {
        "islem_tarihi":           tarih,
        "alim_toplam_nominal":    g(1),
        "satim_toplam_nominal":   g(2),
        "net_nominal":            g(3),
        "gun_basi_nominal":       g(4),
        "gun_sonu_nominal":       g(5),
        "sermaye_orani_gun_basi": g(6),
        "oy_haklari_gun_basi":    g(7),
        "sermaye_orani_gun_sonu": g(8),
        "oy_haklari_gun_sonu":    g(9) if len(row) > 9 else "",
    }


def _map_portfolio_row(row: list) -> dict:
    """
    Portföy şirketi formatı (DenizPortföy vb.):
    İşlem Tarihi | Nitelik | Nominal | Fiyat | İşlem Tutarı |
    Öncesi Nominal | Öncesi Sermaye | Sonrası Nominal | Sonrası Sermaye
    """
    cells = [str(c or "").strip().replace("\n"," ") for c in row]

    # İlk tarih hücresini bul
    tarih_idx = None
    for i, c in enumerate(cells):
        if re.search(r'\d{2}[/\.]\d{2}[/\.]\d{4}', c):
            tarih_idx = i
            break
    if tarih_idx is None:
        return {}

    tarih = cells[tarih_idx].replace("/",".")
    remaining = cells[tarih_idx + 1:]

    # Nitelik bul
    nitelik = ""
    nit_idx = None
    for i, c in enumerate(remaining):
        cl = c.lower()
        if any(k in cl for k in ["alım","alim","alış","alis"]):
            nitelik = "Alım"
            nit_idx = i
            break
        elif any(k in cl for k in ["satım","satim","satış","satis"]):
            nitelik = "Satım"
            nit_idx = i
            break

    # Kalan sayılar
    start = (nit_idx + 1) if nit_idx is not None else 0
    nums = []
    for c in remaining[start:]:
        clean = c.replace(".","").replace(",",".").replace("%","").replace(" ","").strip()
        if clean and re.match(r'^-?[\d.]+$', clean):
            nums.append(c)

    result = {"islem_tarihi": tarih}

    # nominal, fiyat, islem_tutari, oncesi_nominal, oncesi_sermaye, sonrasi_nominal, sonrasi_sermaye
    nominal = nums[0] if len(nums) > 0 else ""
    fiyat   = nums[1] if len(nums) > 1 else ""

    if "Alım" in nitelik:
        result["alim_toplam_nominal"]  = nominal
        result["satim_toplam_nominal"] = "0"
        result["net_nominal"]          = nominal
    elif "Satım" in nitelik:
        result["alim_toplam_nominal"]  = "0"
        result["satim_toplam_nominal"] = nominal
        result["net_nominal"]          = f"-{nominal}" if nominal and not nominal.startswith("-") else nominal
    else:
        result["alim_toplam_nominal"]  = nominal
        result["satim_toplam_nominal"] = "0"
        result["net_nominal"]          = nominal

    if fiyat:
        result["fiyat"] = fiyat
    if len(nums) > 3:
        result["gun_basi_nominal"] = nums[3]
    if len(nums) > 4:
        result["sermaye_orani_gun_basi"] = nums[4]
    if len(nums) > 5:
        result["gun_sonu_nominal"] = nums[5]
    if len(nums) > 6:
        result["sermaye_orani_gun_sonu"] = nums[6]

    return result


def _extract_company(tables: list, text: str) -> str:
    """PDF'den şirket/kişi adını çıkar."""
    # Tablo: "Ad Soyad / Ticaret Ünvanı" : "XXX"
    for table in tables:
        for row in table:
            if not row or len(row) < 2:
                continue
            key = str(row[0] or "").lower()
            val = str(row[1] or "").strip()
            if ("ad soyad" in key or "ticaret" in key or "ünvan" in key) and val:
                # ": XXX" → "XXX"
                val = val.lstrip(": ").strip()
                if val and len(val) > 2:
                    return val

    # Metin: "Bildirime Konu Borsa Şirketi : XXX A.Ş."
    m = re.search(r'Bildirime\s+Konu\s+Borsa\s+Şirketi\s*:\s*(.+?)(?:\n|$)', text)
    if m:
        return m.group(1).strip()

    # "Saygılarımızla" sonrası
    m2 = re.search(
        r'[Ss]aygılarımızla[,.]?\s*\n\s*\n?\s*([A-ZÇĞİÖŞÜ][^\n]{5,}(?:A\.Ş\.|A\.S\.))',
        text
    )
    if m2:
        name = m2.group(1).strip()
        if "KAMUYU" not in name.upper():
            return name

    return ""


def _extract_price(text: str) -> str:
    """Metin içinden fiyat bilgisini çıkar."""
    # "15,06 - 15,10 TL fiyat aralığından" → ortalama hesapla veya aralık ver
    m = re.search(r'([\d.,]+)\s*-\s*([\d.,]+)\s*TL\s*fiyat', text)
    if m:
        return f"{m.group(1)} - {m.group(2)}"

    # "XX,XX TL fiyat"
    m2 = re.search(r'([\d.,]+)\s*TL\s*(?:fiyat|ortalama)', text)
    if m2:
        return m2.group(1)

    # Tablo dışı fiyat
    m3 = re.search(r'[Oo]rtalama[\s]+([0-9.,]+)', text)
    if m3:
        return m3.group(1)

    return ""


def _parse_text_fallback(text: str) -> dict:
    """Tablo çıkarılamadıysa metinden dene."""
    result = {}
    # Tarih
    tm = re.search(r'(\d{2}[/\.]\d{2}[/\.]\d{4})', text)
    if tm:
        result["islem_tarihi"] = tm.group(1).replace("/",".")

    # Alım/Satım nominal
    alim = re.search(r'([\d.]+)\s*TL\s*toplam\s*nominal\s*tutarlı\s*alış', text)
    if alim:
        result["alim_toplam_nominal"] = alim.group(1)

    satim = re.search(r'([\d.]+)\s*TL\s*toplam\s*nominal\s*tutarlı\s*satış', text)
    if satim:
        result["satim_toplam_nominal"] = satim.group(1)

    # Sermaye oranı
    serm = re.findall(r'%\s*([\d.,]+)', text)
    if serm:
        result["sermaye_orani_gun_sonu"] = serm[-1]

    return result if result.get("islem_tarihi") else {}


# ─── DETAY FETCH ──────────────────────────────────────────

def fetch_bildirim_pdf(session, disc_index: str, log_fn=print) -> dict:
    """KAP BildirimPdf API: /tr/api/BildirimPdf/{discIndex}"""
    url = f"{BASE_URL}/tr/api/BildirimPdf/{disc_index}"
    try:
        headers_pdf = {
            "User-Agent": HEADERS["User-Agent"],
            "Accept": "application/pdf,*/*",
            "Accept-Language": "tr",
            "Referer": f"{BASE_URL}/tr/Bildirim/{disc_index}",
        }
        r = session.get(url, headers=headers_pdf, timeout=30)
        log_fn(f"    📄 BildirimPdf: HTTP {r.status_code} ({len(r.content)} bytes)")

        if r.status_code != 200 or len(r.content) < 500:
            return {}

        if b'%PDF' in r.content[:20]:
            return _parse_kap_pdf(r.content, log_fn)
        else:
            # HTML gelmiş olabilir
            return _parse_html_response(r.text, log_fn)

    except Exception as e:
        log_fn(f"    ✗ BildirimPdf hata: {e}")
        return {}


def _parse_html_response(html: str, log_fn=print) -> dict:
    """BildirimPdf HTML dönerse parse et."""
    soup = BeautifulSoup(html, "lxml")
    for table in soup.find_all("table"):
        for row in table.find_all("tr"):
            cells = [td.get_text(strip=True).replace("\n"," ")
                     for td in row.find_all(["td","th"])]
            joined = " ".join(cells)
            if (len(cells) >= 5 and
                re.search(r'\d{2}[/\.]\d{2}[/\.]\d{4}', joined)):
                if len(cells) >= 9:
                    return _map_standard_row(cells)
    return {}


def fetch_html_detail(session, disc_index: str, log_fn=print) -> dict:
    """ÖDA tipi bildirimler için HTML sayfasında tablo."""
    result = {"data_rows": [], "fiyat": "", "sirket_adi": ""}
    try:
        url = f"{BASE_URL}/tr/Bildirim/{disc_index}"
        headers_html = {
            "User-Agent": HEADERS["User-Agent"],
            "Accept": "text/html,application/xhtml+xml",
            "Accept-Language": "tr",
            "Referer": f"{BASE_URL}/tr",
        }
        r = session.get(url, headers=headers_html, timeout=20)
        if r.status_code != 200:
            return result
    except Exception:
        return result

    soup = BeautifulSoup(r.text, "lxml")

    for sel in ["div.comp-name","span.comp-name","a.comp-name"]:
        el = soup.select_one(sel)
        if el:
            name = el.get_text(strip=True)
            if name and "KAMUYU AYDINLATMA" not in name.upper():
                result["sirket_adi"] = name
                break

    for table in soup.find_all("table"):
        for row in reversed(table.find_all("tr")):
            cells = [td.get_text(strip=True).replace("\n"," ")
                     for td in row.find_all(["td","th"])]
            joined = " ".join(cells)
            if (len(cells) >= 5 and
                re.search(r'\d{2}[/\.]\d{2}[/\.]\d{4}', joined) and
                '%' in joined):
                result["data_rows"].append(cells)
        if result["data_rows"]:
            break

    body = soup.get_text(" ", strip=True)
    fm = re.search(r'[Oo]rtalama[\s]+([0-9.,]+)[\s]*fiyat', body)
    if fm:
        result["fiyat"] = fm.group(1)

    return result


# ─── ANA DETAY TOPLAMA ────────────────────────────────────

def fetch_and_enrich(session, disc: dict, log_fn=print) -> dict:
    """Bildirim için detay çek ve satırı zenginleştir."""
    row = dict(disc)
    disc_index = disc.get("disc_index", "")
    if not disc_index:
        return row

    # ── Yöntem 1: HTML (ÖDA tipleri — tablo HTML'de) ──
    html = fetch_html_detail(session, disc_index, log_fn)

    if html.get("sirket_adi"):
        if "KAMUYU AYDINLATMA" not in html["sirket_adi"].upper():
            row["sirket"] = html["sirket_adi"]

    if html.get("data_rows"):
        cells = html["data_rows"][0]
        def g(idx):
            return cells[idx].strip() if idx < len(cells) else ""
        row["islem_tarihi"]           = g(0)
        row["alim_toplam_nominal"]    = g(1)
        row["satim_toplam_nominal"]   = g(2)
        row["net_nominal"]            = g(3)
        row["gun_basi_nominal"]       = g(4)
        row["gun_sonu_nominal"]       = g(5)
        row["sermaye_orani_gun_basi"] = g(6)
        row["oy_haklari_gun_basi"]    = g(7)
        row["sermaye_orani_gun_sonu"] = g(8)
        row["oy_haklari_gun_sonu"]    = g(9)
        if html.get("fiyat"):
            row["fiyat"] = html["fiyat"]
        log_fn(f"    ✓ HTML tablo: {g(0)}")
        return row

    # ── Yöntem 2: BildirimPdf API → PDF parse ──
    pdf = fetch_bildirim_pdf(session, disc_index, log_fn)

    if not pdf or not pdf.get("islem_tarihi"):
        log_fn(f"    ⚠ Veri bulunamadı (HTML + PDF)")
        return row

    # Şirket adı
    if pdf.get("sirket_adi"):
        if "KAMUYU AYDINLATMA" not in pdf["sirket_adi"].upper():
            row["sirket"] = pdf["sirket_adi"]

    # Tüm alanları aktar
    for key in ["islem_tarihi", "alim_toplam_nominal", "satim_toplam_nominal",
                "net_nominal", "gun_basi_nominal", "gun_sonu_nominal",
                "sermaye_orani_gun_basi", "oy_haklari_gun_basi",
                "sermaye_orani_gun_sonu", "oy_haklari_gun_sonu", "fiyat"]:
        if pdf.get(key):
            row[key] = pdf[key]

    return row


def fetch_details_requests(session, disclosures: list, log_fn=print) -> list:
    log_fn(f"  🌐 {len(disclosures)} detay çekiliyor...")
    enriched = []
    for i, disc in enumerate(disclosures, 1):
        ilgili = disc.get("ilgili_sirket", "?")
        sirket = disc.get("sirket", "")[:35]
        log_fn(f"  [{i}/{len(disclosures)}] {ilgili} ← {sirket}")

        for attempt in range(2):
            try:
                row = fetch_and_enrich(session, disc, log_fn)
                if row.get("islem_tarihi"):
                    log_fn(f"    ✓ {row['islem_tarihi']} | "
                           f"Alım:{row.get('alim_toplam_nominal','-')} | "
                           f"Sermaye:{row.get('sermaye_orani_gun_sonu','-')}")
                enriched.append(row)
                break
            except Exception as e:
                if attempt == 0:
                    log_fn(f"    ↻ Tekrar deneniyor...")
                    time.sleep(1)
                else:
                    log_fn(f"    ✗ {e}")
                    enriched.append(dict(disc))

        if i < len(disclosures):
            time.sleep(0.5)

    return enriched


# ─── ANA FONKSİYON ───────────────────────────────────────

def scrape_pay_alim_satim(start_date: date, end_date: date,
                          log_fn=print) -> list:
    log_fn(f"📡 KAP: {start_date.strftime('%d.%m.%Y')} → {end_date.strftime('%d.%m.%Y')}")

    session = make_session()
    log_fn("  ✓ Session kuruldu")

    raw = fetch_main_api(session, start_date, end_date)

    if not raw:
        log_fn("  ⚠ API yanıt vermedi — demo veri")
        return get_demo_data(start_date, end_date)

    log_fn(f"  {len(raw)} kayıt içinde Pay Alım Satım filtreleniyor...")
    filtered    = filter_pay_alim_satim(raw)
    disclosures = [normalize_item(i) for i in filtered]
    log_fn(f"  ✓ {len(disclosures)} bildirim bulundu")

    enriched = fetch_details_requests(session, disclosures, log_fn=log_fn)
    log_fn(f"✅ Tamamlandı: {len(enriched)} bildirim")
    return enriched


def get_demo_data(start_date: date, end_date: date) -> list:
    d1 = start_date.strftime("%d.%m.%Y")
    return [
        {"no":"1583033","tarih":f"{d1} 18:46","kod":"ALNUS, ANC",
         "sirket":"ALNUS YATIRIM MENKUL DEĞERLER A.Ş.",
         "konu":"Pay Alım Satım Bildirimi","ozet":"ISKPL Pay Alım Bildirimi",
         "link":f"{BASE_URL}/tr/Bildirim/1583033","disc_index":"1583033",
         "islem_tarihi":"31.03.2026","alim_toplam_nominal":"93.925.229",
         "satim_toplam_nominal":"10.259","net_nominal":"93.914.970",
         "gun_basi_nominal":"0","gun_sonu_nominal":"93.914.970",
         "sermaye_orani_gun_basi":"% 0","oy_haklari_gun_basi":"% 0",
         "sermaye_orani_gun_sonu":"% 6,26","oy_haklari_gun_sonu":"% 6,26",
         "ilgili_sirket":"ISKPL","fiyat":"12,50"},
    ]


# ─── EXCEL ───────────────────────────────────────────────

COLUMNS_MAP = {
    "no":"No", "tarih":"Yayın Tarihi", "kod":"Hisse Kodu",
    "sirket":"Aracı Kurum", "konu":"Konu", "ozet":"Özet",
    "ilgili_sirket":"İlgili Şirket", "islem_tarihi":"İşlem Tarihi",
    "fiyat":"Ort. Fiyat (TL)", "alim_toplam_nominal":"Alım Nominal (TL)",
    "satim_toplam_nominal":"Satım Nominal (TL)", "net_nominal":"Net Nominal (TL)",
    "gun_basi_nominal":"Gün Başı Nominal (TL)", "gun_sonu_nominal":"Gün Sonu Nominal (TL)",
    "sermaye_orani_gun_basi":"Sermaye Oranı Gün Başı (%)",
    "oy_haklari_gun_basi":"Oy Hakları Gün Başı (%)",
    "sermaye_orani_gun_sonu":"Sermaye Oranı Gün Sonu (%)",
    "oy_haklari_gun_sonu":"Oy Hakları Gün Sonu (%)",
    "link":"KAP Linki",
}

COL_WIDTHS = {
    "No":10,"Yayın Tarihi":18,"Hisse Kodu":12,"Aracı Kurum":32,
    "Konu":36,"Özet":28,"İlgili Şirket":14,"İşlem Tarihi":14,
    "Ort. Fiyat (TL)":14,"Alım Nominal (TL)":20,"Satım Nominal (TL)":20,
    "Net Nominal (TL)":20,"Gün Başı Nominal (TL)":22,"Gün Sonu Nominal (TL)":22,
    "Sermaye Oranı Gün Başı (%)":24,"Oy Hakları Gün Başı (%)":22,
    "Sermaye Oranı Gün Sonu (%)":24,"Oy Hakları Gün Sonu (%)":22,"KAP Linki":20,
}


def save_to_excel(enriched: list, start_date: date, end_date: date,
                  output_dir: str = ".") -> tuple:
    os.makedirs(output_dir, exist_ok=True)
    fname    = f"KAP_PayAlimSatim_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, fname)

    rows = [{col: item.get(key,"") for key,col in COLUMNS_MAP.items()} for item in enriched]
    df   = pd.DataFrame(rows)
    df.to_excel(filepath, index=False, sheet_name="Pay Alım Satım")

    wb  = load_workbook(filepath)
    ws  = wb.active
    hf  = PatternFill("solid", start_color="1F4E79")
    af  = PatternFill("solid", start_color="D6E4F0")
    wf  = PatternFill("solid", start_color="FFFFFF")
    t   = Side(style="thin", color="BDD7EE")
    brd = Border(left=t, right=t, top=t, bottom=t)
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    bfont = Font(name="Arial", size=9)
    lfont = Font(name="Arial", size=9, color="0563C1", underline="single")

    for cell in ws[1]:
        cell.fill = hf; cell.font = hfont; cell.border = brd
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 38

    lc = next((i for i,c in enumerate(ws[1],1) if c.value=="KAP Linki"), None)
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = af if i%2==0 else wf
        for cell in row:
            cell.fill = fill; cell.border = brd
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if lc and cell.column==lc and cell.value:
                cell.font = lfont; cell.hyperlink = str(cell.value)
                cell.value = "Bildirimi Görüntüle"
            else:
                cell.font = bfont

    for idx, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = COL_WIDTHS.get(col_name, 15)
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Özet")
    ws2["A1"] = "KAP Pay Alım Satım Bildirimleri"
    ws2["A1"].font = Font(bold=True, size=14, name="Arial", color="1F4E79")
    for r,(lbl,val) in enumerate([
        ("Başlangıç", start_date.strftime("%d.%m.%Y")),
        ("Bitiş",     end_date.strftime("%d.%m.%Y")),
        ("Kayıt",     str(len(enriched))),
        ("Rapor",     datetime.now().strftime("%d.%m.%Y %H:%M")),
    ], start=3):
        ws2.cell(r,1,lbl).font = Font(bold=True, name="Arial", size=10)
        ws2.cell(r,2,val).font = Font(name="Arial", size=10)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 22
    wb.save(filepath)
    print(f"✓ Excel: {filepath}")
    return filepath, df


if __name__ == "__main__":
    today = date.today()
    if len(sys.argv) >= 3:
        try:
            start = datetime.strptime(sys.argv[1], "%d.%m.%Y").date()
            end   = datetime.strptime(sys.argv[2], "%d.%m.%Y").date()
        except ValueError:
            print("Kullanım: python kap_scraper.py GG.AA.YYYY GG.AA.YYYY [klasör]")
            sys.exit(1)
    else:
        start = today; end = today

    out = sys.argv[3] if len(sys.argv) >= 4 else "."
    data = scrape_pay_alim_satim(start, end)
    fp, df = save_to_excel(data, start, end, out)
    print(f"✓ {len(df)} kayıt → {fp}")
